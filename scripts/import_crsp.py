"""Import the KRA CRSP workbook into the versioned database.

Usage: python3 scripts/import_crsp.py <xlsx-path> <dataset-name> <effective-date>
Reads the workbook read-only, normalizes fields, writes CSVs and loads them with psql.
Questionable records are flagged in data_validation_issues, never silently fixed.
"""

import csv
import json
import os
import re
import subprocess
import sys
import uuid

import openpyxl

FUEL_MAP = {
    "GASOLINE": "PETROL",
    "PETROL": "PETROL",
    "DIESEL": "DIESEL",
    "DEISEL": "DIESEL",
    "DIESEL ": "DIESEL",
    "DI ESEL": "DIESEL",
    "DIESE L": "DIESEL",
    "ELECTRIC": "ELECTRIC",
    "ELECCTRIC": "ELECTRIC",
    "EELCTRIC": "ELECTRIC",
    "ELECTRIC(EV)": "ELECTRIC",
    "HYBRID": "HYBRID",
    "PLUG-IN HYBRID": "PLUG_IN_HYBRID",
    "PLUG-IN-HYBRID": "PLUG_IN_HYBRID",
    "PETROL/ELECTRIC": "PETROL_ELECTRIC",
    "PETROL/DIESEL": "PETROL_DIESEL",
    "CNG": "CNG",
    "LNG": "LNG",
}


def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9]+", " ", str(s).upper())).strip()


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_fuel(raw):
    if raw is None:
        return None, "missing_fuel"
    key = re.sub(r"\s+", " ", str(raw).strip().upper())
    if key in FUEL_MAP:
        return FUEL_MAP[key], None
    compact = key.replace(" ", "")
    for k, v in FUEL_MAP.items():
        if k.replace(" ", "") == compact:
            return v, "fuel_formatting_inconsistency"
    return None, "unrecognized_fuel"


def parse_cc(raw):
    if raw is None:
        return None, "missing_engine_capacity"
    s = str(raw).strip()
    if re.fullmatch(r"\d+(\.0+)?", s):
        return int(float(s)), None
    m = re.fullmatch(r"(\d+)\s*CC", s.upper())
    if m:
        return int(m.group(1)), "engine_capacity_formatting"
    return None, "non_numeric_engine_capacity"


def parse_num(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def main():
    path, name, effective = sys.argv[1], sys.argv[2], sys.argv[3]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    dataset_id = str(uuid.uuid4())
    issues = []
    out = "/tmp/crsp_import"
    os.makedirs(out, exist_ok=True)

    def issue(table, rid, row, itype, severity, detail):
        issues.append([str(uuid.uuid4()), dataset_id, table, rid, row, itype, severity, detail])

    # ---- vehicles ----
    ws = wb[[s for s in sheets if s.startswith("M.Vehicle")][0]]
    headers = ["Make", "Model", "Model number", "Transmission", "Drive Configuration",
               "Engine Capacity", "Body Type", "GVW", "Seating", "Fuel", "CRSP (KES.)"]
    vrows, seen = [], {}
    for i, r in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        r = list(r) + [None] * 11
        if not any(x is not None and str(x).strip() != "" for x in r[:11]):
            continue
        vid = str(uuid.uuid4())
        make, model, modelno = clean(r[0]), clean(r[1]), clean(r[2])
        fuel, fissue = parse_fuel(clean(r[9]))
        cc, cissue = parse_cc(clean(r[5]))
        crsp = parse_num(r[10])
        flags = [f for f in (fissue, cissue) if f]
        if crsp is None:
            flags.append("missing_crsp")
            issue("vehicles", vid, i, "missing_crsp", "critical", "No CRSP value in source row")
        if not make or not model:
            flags.append("missing_required_field")
            issue("vehicles", vid, i, "missing_required_field", "critical", "Make or Model empty")
        if fissue:
            issue("vehicles", vid, i, fissue, "warning", f"Fuel value: {r[9]!r}")
        if cissue:
            issue("vehicles", vid, i, cissue, "warning", f"Engine capacity value: {r[5]!r}")
        key = (norm(make), norm(model), norm(modelno), norm(r[5]), fuel)
        if key in seen:
            flags.append("duplicate_variant")
            issue("vehicles", vid, i, "duplicate_variant", "warning",
                  f"Same make/model/model-number/engine/fuel as source row {seen[key]}")
        else:
            seen[key] = i
        search = norm(" ".join(filter(None, [make, model, modelno, clean(r[6]), clean(r[3]),
                                             clean(r[4]), clean(r[9]), clean(r[5])])))
        original = {h: (str(v) if v is not None else None) for h, v in zip(headers, r[:11])}
        vrows.append([vid, dataset_id, i, make, model, modelno, clean(r[3]), clean(r[4]),
                      clean(r[5]), cc, clean(r[6]), clean(r[7]), clean(r[8]), clean(r[9]),
                      fuel, crsp, json.dumps(original), search, "{" + ",".join(flags) + "}"])

    # ---- motorcycles ----
    ws = wb[[s for s in sheets if s.startswith("Motor Cycles")][0]]
    mheaders = ["Make", "Model", "Model number", "Transmission", "Engine Capacity",
                "seating", "Fuel", "CRSP (KES)"]
    mrows, mseen = [], {}
    for i, r in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        r = list(r) + [None] * 8
        if not any(x is not None and str(x).strip() != "" for x in r[:8]):
            continue
        mid = str(uuid.uuid4())
        make, model, modelno = clean(r[0]), clean(r[1]), clean(r[2])
        fuel, fissue = parse_fuel(clean(r[6]))
        cc, cissue = parse_cc(clean(r[4]))
        crsp = parse_num(r[7])
        flags = [f for f in (fissue, cissue) if f]
        if crsp is None:
            flags.append("missing_crsp")
            issue("motorcycles", mid, i, "missing_crsp", "critical", "No CRSP value in source row")
        if fissue:
            issue("motorcycles", mid, i, fissue, "warning", f"Fuel value: {r[6]!r}")
        if cissue:
            issue("motorcycles", mid, i, cissue, "warning", f"Engine capacity value: {r[4]!r}")
        key = (norm(make), norm(model), norm(modelno), norm(r[4]), fuel)
        if key in mseen:
            flags.append("duplicate_variant")
            issue("motorcycles", mid, i, "duplicate_variant", "warning",
                  f"Duplicate of source row {mseen[key]}")
        else:
            mseen[key] = i
        search = norm(" ".join(filter(None, [make, model, modelno, clean(r[3]), clean(r[6]), clean(r[4])])))
        original = {h: (str(v) if v is not None else None) for h, v in zip(mheaders, r[:8])}
        mrows.append([mid, dataset_id, i, make, model, modelno, clean(r[3]), clean(r[4]), cc,
                      clean(r[5]), clean(r[6]), fuel, crsp, json.dumps(original), search,
                      "{" + ",".join(flags) + "}"])

    # ---- machinery (brand header rows followed by model rows) ----
    ws = wb[[s for s in sheets if s.startswith("Tractors")][0]]
    krows, brand = [], None
    for i, r in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        r = list(r) + [None] * 6
        model, rating, price = clean(r[0]), clean(r[1]), parse_num(r[2])
        if model and rating is None and price is None:
            if model.upper() != "KSHS":
                brand = model
            continue
        if model is None and rating is None and price is None:
            continue
        kid = str(uuid.uuid4())
        flags = []
        if price is None:
            flags.append("missing_crsp")
            issue("machinery", kid, i, "missing_crsp", "critical", "No price in source row")
        if brand is None:
            flags.append("missing_make")
            issue("machinery", kid, i, "missing_make", "warning", "No brand header above this row")
        search = norm(" ".join(filter(None, [brand, model, rating])))
        original = {"MODEL": model, "HORSEPOWER/CC/KW": rating, "KSHS": str(r[2]) if r[2] is not None else None,
                    "BRAND_HEADER": brand}
        krows.append([kid, dataset_id, i, brand, model, rating, price, json.dumps(original),
                      search, "{" + ",".join(flags) + "}"])

    def dump(fname, rows):
        p = os.path.join(out, fname)
        with open(p, "w", newline="") as f:
            csv.writer(f).writerows(rows)
        return p

    vp, mp, kp, ip = (dump("vehicles.csv", vrows), dump("motorcycles.csv", mrows),
                      dump("machinery.csv", krows), dump("issues.csv", issues))

    def psql(sql, stdin=None):
        res = subprocess.run(["psql", "-v", "ON_ERROR_STOP=1", "-c", sql],
                             input=stdin, capture_output=True, text=True)
        if res.returncode != 0:
            raise SystemExit(res.stderr)
        return res.stdout

    psql(
        "INSERT INTO vehicle_datasets (id,name,source_file,effective_date,status,notes) VALUES "
        f"('{dataset_id}','{name}','{os.path.basename(path)}','{effective}','draft',"
        "'Imported from the KRA CRSP workbook. Internal worksheet names read July 2025 - REQUIRES VERIFICATION.')"
    )
    for p, table, cols in [
        (vp, "vehicles", "id,dataset_id,source_row,make,model,model_number,transmission,drive_configuration,engine_capacity_raw,engine_capacity_cc,body_type,gvw,seating,fuel_raw,fuel_normalized,crsp_kes,original_row_data,search_text,flags"),
        (mp, "motorcycles", "id,dataset_id,source_row,make,model,model_number,transmission,engine_capacity_raw,engine_capacity_cc,seating,fuel_raw,fuel_normalized,crsp_kes,original_row_data,search_text,flags"),
        (kp, "machinery", "id,dataset_id,source_row,make,model,rating_raw,crsp_kes,original_row_data,search_text,flags"),
        (ip, "data_validation_issues", "id,dataset_id,record_table,record_id,source_row,issue_type,severity,detail"),
    ]:
        with open(p) as f:
            psql(f"COPY {table} ({cols}) FROM STDIN WITH (FORMAT csv)", stdin=f.read())

    print(json.dumps({"dataset_id": dataset_id, "vehicles": len(vrows), "motorcycles": len(mrows),
                      "machinery": len(krows), "issues": len(issues)}, indent=2))


if __name__ == "__main__":
    main()
